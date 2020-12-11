from pages.datadriven import DataDriven
from pages.page_objects import *
import openpyxl


book = openpyxl.load_workbook(r'C:\Users\GT Sanatorium\PycharmProjects\Ecommerce\Testdata\data.xlsx')
sheet = book.active
rows = sheet.max_row
columns = sheet.max_column


class TestDataDriven:
    def test_ddt(self, browser):
        self.driver = browser
        self.c = DataDriven(self.driver)
        self.c.home_page()
        self.c.cursor_action(registration)
        self.c.button(sign_up)
        for i in range(2,rows+1):
            self.c.text_box(username,sheet.cell(i, 1).value)
            self.c.text_box(mobile_no,sheet.cell(i, 2).value)
            self.c.text_box(email,sheet.cell(i, 3).value)
            self.c.text_box(password,sheet.cell(i,4).value)
        self.c.exit_browser()

